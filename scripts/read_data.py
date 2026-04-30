import csv
from openpyxl import load_workbook

def clean_tin(tin):
    return "".join(ch for ch in str(tin) if ch.isdigit())


COMPANY_INFO_FILE = "input/company_info.csv"
RAW_TAX_DATA = "input/raw_data.xlsx"

def run():
    # ---- READ COMPANY INFO CSV ----
    company_info = {}

    with open(COMPANY_INFO_FILE, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)

        for record in reader:
            tin = record.get("TIN")
            if not tin:
                continue

            key = clean_tin(tin)

            company_info[key] = {
                "ADDRESS": record.get("ADDRESS") or "",
                "ZIP": record.get("ZIP") or "",
                "ATC": record.get("ATC") or ""
            }

    # ---- READ RAW MONTHLY DATA FILE ----
    data_wb = load_workbook(RAW_TAX_DATA, data_only=True)

    month_sheets = data_wb.sheetnames

    companies = {}

    for sheet_name in month_sheets:
        ws = data_wb[sheet_name]

        headers = [cell.value for cell in ws[3]]

        for row in ws.iter_rows(min_row=4, values_only=True):
            record = dict(zip(headers, row))

            company_name = record.get("COMPANY_NAME")
            tin = record.get("TIN")
            net = record.get("NET") or 0

            if not company_name or not tin:
                continue

            key = clean_tin(tin)

            if key not in companies:
                companies[key] = {
                    "COMPANY_NAME": company_name,
                    "TIN": tin,
                    "ADDRESS": company_info.get(key, {}).get("ADDRESS", ""),
                    "ZIP": company_info.get(key, {}).get("ZIP", ""),
                    "MONTHS": {month: 0 for month in month_sheets},
                    "ATC": company_info.get(key, {}).get("ATC", "")
                }

            companies[key]["MONTHS"][sheet_name] = net

    cleaned_rows = []

    for company in companies.values():
        total = sum(company["MONTHS"].values())
        tax_withheld = total * 0.01

        row = [
            company["COMPANY_NAME"],
            company["TIN"],
            company["ADDRESS"],
            company["ZIP"],
            company["ATC"]
        ]

        for month in month_sheets:
            row.append(company["MONTHS"][month])

        row.append(total)
        row.append(tax_withheld)

        cleaned_rows.append(row)

    cleaned_rows.sort(key=lambda x: x[0])

    headers = ["COMPANY_NAME", "TIN", "ADDRESS", "ZIP", "ATC"] + month_sheets + ["TOTAL", "TAX_WITHHELD"]

    with open("output/cleaned_data.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(cleaned_rows)

    print("Created output/cleaned_data.csv")

if __name__ == "__main__":
    run()